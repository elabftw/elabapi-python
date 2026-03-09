"""
eLabFTW Resource Export Script
--------------------------------

This script connects to an eLabFTW instance via the API and exports all items
from a specific category.

For each item, the script:
- Extracts the internal eLabFTW resource ID
- Extracts the item title
- Reads all configured extra fields (custom metadata fields)
- Exports the data in a predefined column order

The export is written to:
- A CSV file (for universal compatibility)
- An XLSX Excel file (with formatting, filters, and auto column width)

The FIELD_ORDER list defines the exact column structure of the export.
This ensures a stable and reproducible layout, which is important for:
- Regulatory documentation
- Data imports into other systems
- Standardized reporting
"""

import elabapi_python
import json
import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

#########################
#        CONFIG         #
#########################

# Base URL of your eLabFTW API (must include /api/v2)
API_HOST_URL = 'https://YOUR-URL/api/v2'

# Personal API key generated in eLabFTW
API_KEY = 'YOUR API Key'

# Category ID from which items should be exported
CATEGORY_ID = 123456

# Output filenames
OUTPUT_FILE = "export.csv"
OUTPUT_XLSX = "export.xlsx"

#########################
#   API CONFIGURATION   #
#########################

# Create configuration object for the eLabFTW API
configuration = elabapi_python.Configuration()
configuration.api_key['api_key'] = API_KEY
configuration.api_key_prefix['api_key'] = 'Authorization'
configuration.host = API_HOST_URL

configuration.debug = False

# Set to True if valid SSL certificates are used
configuration.verify_ssl = True

# Create API client
api_client = elabapi_python.ApiClient(configuration)

# Set authorization header manually (required for authentication)
api_client.set_default_header(
    header_name='Authorization',
    header_value=API_KEY
)

# Load Items API endpoint
itemsApi = elabapi_python.ItemsApi(api_client)

###############################
#   DEFINE EXPORT ORDER HERE  #
###############################

"""
FIELD_ORDER defines the exact column order of metadata fields
in the exported files.

Why is this necessary?

eLabFTW stores extra fields dynamically in JSON format.
Without explicitly defining the order:
- Column positions could change
- Fields might appear in random order
- Downstream processing (e.g., Excel templates, validation scripts)
  could break

By defining FIELD_ORDER:
- The export structure remains stable
- Reports remain consistent
- Future modifications can be controlled centrally

You can modify this list to match your own category setup.
Simply replace the example field names below with the exact
field titles used in your system.
"""

FIELD_ORDER = [
    "Record Number",            # e.g. internal running number
    "Project ID",               # e.g. reference or file number
    "Organism Name",            # e.g. E. coli strain
    "Gene / Target",            # e.g. transgene or modification
    "Storage Location",         # e.g. freezer or storage unit
    "Experiment Purpose",       # short description of use
    "Donor Organism",
    "Recipient Organism",
    "Vector",
    "Resistance Marker",
    "Sequence Information",
    "Risk Assessment Reference",
    "Created By",
    "Created At",
    "Comments"
]

# --------------------------------------------------
# ----------- SCRIPT STARTS HERE ------------------
# --------------------------------------------------

print("Starting export...")

# Retrieve all items from the specified category
items = itemsApi.read_items(cat=CATEGORY_ID)

print(f"Items found: {len(items)}")

rows = []

# Iterate through all retrieved items
for item in items:

    row = {}

    # Internal eLabFTW resource ID (primary identifier)
    row["Ressourcen ID"] = item.id or ""

    # Item title
    row["Titel"] = item.title or ""

    # Metadata is stored as JSON string
    metadata_raw = item.metadata

    if metadata_raw:
        metadata = json.loads(metadata_raw)
        extra_fields = metadata.get("extra_fields", {})
    else:
        extra_fields = {}

    normalized_fields = {k.strip().lower(): v for k, v in extra_fields.items()}

    for field in FIELD_ORDER:
        key = field.strip().lower()
        if key in normalized_fields:
            value = normalized_fields[key].get("value", "")
        else:
            value = ""

        # Convert list-type fields into comma-separated string
        if isinstance(value, list):
            value = ", ".join(value)

        row[field] = value

    rows.append(row)

#########################
#       WRITE CSV       #
#########################

# Define final column structure
csv_columns = ["Ressourcen ID", "Titel"] + FIELD_ORDER

# Write CSV file (UTF-8 encoding for special characters)
with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=csv_columns)
    writer.writeheader()
    writer.writerows(rows)

print("CSV export finished.")

#########################
#      WRITE XLSX       #
#########################

wb = Workbook()
ws = wb.active
ws.title = "Export"

# Write header row
ws.append(csv_columns)

# Make header row bold
for cell in ws[1]:
    cell.font = Font(bold=True)

# Write data rows
for row in rows:
    ws.append([row[col] for col in csv_columns])

# Enable auto-filter for the entire sheet
ws.auto_filter.ref = ws.dimensions

# Automatically adjust column width based on content length
for col in ws.columns:
    max_length = 0
    column = col[0].column
    column_letter = get_column_letter(column)

    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass

    adjusted_width = max_length + 2
    ws.column_dimensions[column_letter].width = adjusted_width

# Save Excel file
wb.save(OUTPUT_XLSX)

print("XLSX export finished successfully.")
